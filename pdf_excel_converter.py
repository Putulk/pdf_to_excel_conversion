import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
from pathlib import Path
import os
from PIL import Image
import pytesseract
from pdf2image import convert_from_path

class AutoHotelPDFConverter:
    def __init__(self, folder_path):
        self.folder_path = Path(folder_path)
        self.hotels_data = []

    def extract_candlewood_data(self, pdf_path):
        """Extract data from Candlewood Burlington format PDF"""
        data = {
            'name': 'Candlewood Burlington',
            'for_day': {},
            'mtd': {},
            'ytd': {}
        }

        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()
            lines = text.split('\n')

            for line in lines:
                if 'Rooms Occupied' in line and 'minus Comp' not in line and '%' not in line:
                    parts = re.findall(r'\d+', line)
                    if len(parts) >= 3:
                        data['for_day']['rooms_sold'] = int(parts[0])
                        data['mtd']['rooms_sold'] = int(parts[1])
                        data['ytd']['rooms_sold'] = int(parts[2])

                if 'Total Rooms in Hotel' in line:
                    parts = re.findall(r'\d+', line)
                    if len(parts) >= 3:
                        data['for_day']['total_rooms'] = int(parts[0])
                        data['mtd']['total_rooms'] = int(parts[1])
                        data['ytd']['total_rooms'] = int(parts[2])

                if 'Rooms Occupied minus Comp' in line:
                    parts = re.findall(r'\d+', line)
                    if len(parts) >= 6:
                        data['for_day']['comp_rooms'] = int(parts[0]) - int(parts[3])
                        data['mtd']['comp_rooms'] = int(parts[1]) - int(parts[4])
                        data['ytd']['comp_rooms'] = int(parts[2]) - int(parts[5])

                if 'Out of Order Rooms' in line:
                    parts = re.findall(r'\d+', line)
                    if len(parts) >= 3:
                        data['for_day']['ooo_rooms'] = int(parts[0])
                        data['mtd']['ooo_rooms'] = int(parts[1])
                        data['ytd']['ooo_rooms'] = int(parts[2])

                if 'ADR minus Comp' in line or ('ADR' in line and 'Revenue' not in line and 'minus' in line):
                    parts = re.findall(r'\d+\.?\d*', line)
                    if len(parts) >= 3:
                        data['for_day']['adr'] = round(float(parts[0]), 2)
                        data['mtd']['adr'] = round(float(parts[1]), 2)
                        data['ytd']['adr'] = round(float(parts[2]), 2)

                if 'RevPar' in line:
                    parts = re.findall(r'\d+\.?\d*', line)
                    if len(parts) >= 3:
                        data['for_day']['revpar'] = round(float(parts[0]), 2)
                        data['mtd']['revpar'] = round(float(parts[1]), 2)
                        data['ytd']['revpar'] = round(float(parts[2]), 2)

                if '% Rooms Occupied' in line and 'minus' not in line:
                    parts = re.findall(r'\d+\.?\d*', line)
                    if len(parts) >= 3:
                        data['for_day']['occp_pct'] = round(float(parts[0]), 2)
                        data['mtd']['occp_pct'] = round(float(parts[1]), 2)
                        data['ytd']['occp_pct'] = round(float(parts[2]), 2)

                if 'Room Revenue' in line and 'Average' not in line and 'Block' not in line and 'Individual' not in line:
                    parts = re.findall(r'\d+[\d,]*\.?\d*', line)
                    if len(parts) >= 3:
                        data['for_day']['room_revenue'] = round(float(parts[0].replace(',', '')), 2)
                        data['mtd']['room_revenue'] = round(float(parts[1].replace(',', '')), 2)
                        data['ytd']['room_revenue'] = round(float(parts[2].replace(',', '')), 2)

                if 'Food And Beverage Revenue' in line or 'F & B Revenue' in line:
                    parts = re.findall(r'\d+[\d,]*\.?\d*', line)
                    if len(parts) >= 3:
                        data['for_day']['fb_revenue'] = round(float(parts[0].replace(',', '')), 2)
                        data['mtd']['fb_revenue'] = round(float(parts[1].replace(',', '')), 2)
                        data['ytd']['fb_revenue'] = round(float(parts[2].replace(',', '')), 2)

                if 'Other Revenue' in line:
                    parts = re.findall(r'\d+[\d,]*\.?\d*', line)
                    if len(parts) >= 3:
                        data['for_day']['other_revenue'] = round(float(parts[0].replace(',', '')), 2)
                        data['mtd']['other_revenue'] = round(float(parts[1].replace(',', '')), 2)
                        data['ytd']['other_revenue'] = round(float(parts[2].replace(',', '')), 2)

                if 'Total Revenue' in line:
                    parts = re.findall(r'\d+[\d,]*\.?\d*', line)
                    if len(parts) >= 3:
                        data['for_day']['total_revenue'] = round(float(parts[0].replace(',', '')), 2)
                        data['mtd']['total_revenue'] = round(float(parts[1].replace(',', '')), 2)
                        data['ytd']['total_revenue'] = round(float(parts[2].replace(',', '')), 2)

        return data

    def extract_tps_niagara_data(self, pdf_path):
        """Extract data from TPS Niagara format PDF (supports both text-based and image-based PDFs)"""
        data = {
            'name': 'TPS Niagara',
            'for_day': {},
            'mtd': {},
            'ytd': {}
        }

        # First, try text extraction
        text = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
        except:
            pass

        # If no text found or very short, use OCR
        if not text or len(text.strip()) < 100:
            print("   [INFO] PDF appears to be image-based, using OCR...")
            try:
                # Convert PDF to image
                images = convert_from_path(pdf_path, first_page=1, last_page=1, dpi=300)
                if images:
                    # Perform OCR on the first page
                    text = pytesseract.image_to_string(images[0], config='--psm 6')
                    print("   [INFO] OCR extraction completed")
            except Exception as e:
                print(f"   [ERROR] OCR failed: {e}")
                print("   [INFO] Make sure tesseract is installed:")
                print("          Mac: brew install tesseract")
                print("          Linux: sudo apt-get install tesseract-ocr")
                print("          Windows: Download from https://github.com/UB-Mannheim/tesseract/wiki")
                return data

        if not text:
            print("   [ERROR] Could not extract any text from PDF")
            return data

        lines = text.split('\n')

        print("\n   [DEBUG] Extracting TPS Niagara data...")
        print("   [DEBUG] Column structure: TODAY'S ACTUAL | TODAY'S BUDGET | PTD'S ACTUAL | PTD'S BUDGET | YTD'S ACTUAL | YTD'S BUDGET")

        for line in lines:
            # Extract all numbers (including decimals and commas)
            nums = re.findall(r'\d+(?:,\d+)*\.?\d*-?', line)
            nums = [n.replace(',', '').replace('-', '') for n in nums]

            # TOTAL ROOM SALES
            if 'TOTAL' in line.upper() and 'ROOM' in line.upper() and 'SALES' in line.upper():
                if len(nums) >= 6:
                    try:
                        data['for_day']['room_revenue'] = float(nums[0])
                        data['mtd']['room_revenue'] = float(nums[2])
                        data['ytd']['room_revenue'] = float(nums[4])
                        print(f"   ✓ Room Revenue: Day={nums[0]} | PTD={nums[2]} | YTD={nums[4]}")
                    except Exception as e:
                        print(f"   ✗ Error extracting Room Revenue: {e}")

            # TOTAL F. & B. SALES
            if 'TOTAL' in line.upper() and 'F' in line.upper() and 'B' in line.upper() and 'SALES' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['fb_revenue'] = float(nums[0])
                        data['mtd']['fb_revenue'] = float(nums[2])
                        data['ytd']['fb_revenue'] = float(nums[4])
                        print(f"   ✓ F&B Revenue: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # TOTAL MISC. SALES
            if 'TOTAL' in line.upper() and 'MISC' in line.upper() and 'SALES' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['other_revenue'] = float(nums[0])
                        data['mtd']['other_revenue'] = float(nums[2])
                        data['ytd']['other_revenue'] = float(nums[4])
                        print(f"   ✓ Other Revenue: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # GROSS HOTEL SALES
            if 'GROSS' in line.upper() and 'HOTEL' in line.upper() and 'SALES' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['total_revenue'] = float(nums[0])
                        data['mtd']['total_revenue'] = float(nums[2])
                        data['ytd']['total_revenue'] = float(nums[4])
                        print(f"   ✓ Total Revenue: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # # ROOMS OCCUPIED/SOLD
            if '#' in line and 'ROOMS' in line.upper() and ('OCCUPIED' in line.upper() or 'SOLD' in line.upper()):
                if len(nums) >= 5:
                    try:
                        data['for_day']['rooms_sold'] = int(nums[0])
                        data['mtd']['rooms_sold'] = int(nums[2])
                        data['ytd']['rooms_sold'] = int(nums[4])
                        print(f"   ✓ Rooms Sold: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # # TOTAL ROOMS
            if '#' in line and 'TOTAL' in line.upper() and 'ROOMS' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['total_rooms'] = int(nums[0])
                        data['mtd']['total_rooms'] = int(nums[2])
                        data['ytd']['total_rooms'] = int(nums[4])
                        print(f"   ✓ Total Rooms: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # # OUT OF ORDER
            if '#' in line and 'OUT' in line.upper() and 'OF' in line.upper() and 'ORDER' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['ooo_rooms'] = int(nums[0])
                        data['mtd']['ooo_rooms'] = int(nums[2])
                        data['ytd']['ooo_rooms'] = int(nums[4])
                        print(f"   ✓ OOO Rooms: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # # COMPLIMENTARY ROOMS
            if '#' in line and 'COMPLIMENTARY' in line.upper() and 'ROOMS' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['comp_rooms'] = int(nums[0])
                        data['mtd']['comp_rooms'] = int(nums[2])
                        data['ytd']['comp_rooms'] = int(nums[4])
                        print(f"   ✓ Comp Rooms: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # AVG RATE PER ROOM
            if 'AVG' in line.upper() and 'RATE' in line.upper() and 'ROOM' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['adr'] = float(nums[0])
                        data['mtd']['adr'] = float(nums[2])
                        data['ytd']['adr'] = float(nums[4])
                        print(f"   ✓ ADR: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # OCCUPANCY PCT
            if 'OCCUPANCY' in line.upper() and 'PCT' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['occp_pct'] = float(nums[0])
                        data['mtd']['occp_pct'] = float(nums[2])
                        data['ytd']['occp_pct'] = float(nums[4])
                        print(f"   ✓ Occupancy %: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

            # REV PAR
            if 'REV' in line.upper() and 'PAR' in line.upper():
                if len(nums) >= 5:
                    try:
                        data['for_day']['revpar'] = float(nums[0])
                        data['mtd']['revpar'] = float(nums[2])
                        data['ytd']['revpar'] = float(nums[4])
                        print(f"   ✓ RevPar: {nums[0]} | {nums[2]} | {nums[4]}")
                    except:
                        pass

        # CALCULATE MISSING VALUES
        print("\n   [DEBUG] Calculating missing values...")

        for period in ['for_day', 'mtd', 'ytd']:
            period_data = data[period]

            # Calculate ADR if missing: ADR = Room Revenue / Rooms Sold
            if not period_data.get('adr') and period_data.get('room_revenue') and period_data.get('rooms_sold'):
                if period_data['rooms_sold'] > 0:
                    period_data['adr'] = round(period_data['room_revenue'] / period_data['rooms_sold'], 2)
                    print(f"   ✓ Calculated {period} ADR: {period_data['adr']}")

            # Calculate Occupancy % if missing
            if not period_data.get('occp_pct') and period_data.get('rooms_sold') and period_data.get('total_rooms'):
                if period_data['total_rooms'] > 0:
                    period_data['occp_pct'] = round((period_data['rooms_sold'] / period_data['total_rooms']) * 100, 2)
                    print(f"   ✓ Calculated {period} Occupancy %: {period_data['occp_pct']}")

            # Calculate RevPar if missing
            if not period_data.get('revpar') and period_data.get('room_revenue') and period_data.get('total_rooms'):
                if period_data['total_rooms'] > 0:
                    period_data['revpar'] = round(period_data['room_revenue'] / period_data['total_rooms'], 2)
                    print(f"   ✓ Calculated {period} RevPar: {period_data['revpar']}")

            # Calculate Total Revenue if missing
            if not period_data.get('total_revenue'):
                room = period_data.get('room_revenue', 0)
                fb = period_data.get('fb_revenue', 0)
                other = period_data.get('other_revenue', 0)
                if room or fb or other:
                    period_data['total_revenue'] = round(room + fb + other, 2)
                    print(f"   ✓ Calculated {period} Total Revenue: {period_data['total_revenue']}")

        print(f"\n   [DEBUG] Final extracted data for TPS Niagara:")
        print(f"      For Day: {data['for_day']}")
        print(f"      MTD: {data['mtd']}")
        print(f"      YTD: {data['ytd']}")

        return data

    def extract_bayview_data(self, pdf_path):
        """Extract data from Bayview Wildwood format PDF"""
        data = {
            'name': 'Bayview Wildwood',
            'for_day': {},
            'mtd': {},
            'ytd': {}
        }

        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()
            lines = text.split('\n')

            for line in lines:
                if line.strip().startswith('Total Rooms') and 'Revenue' not in line and 'Occupied' not in line:
                    parts = re.findall(r'\d+', line)
                    if len(parts) >= 5:
                        data['for_day']['total_rooms'] = int(parts[0])
                        data['mtd']['total_rooms'] = int(parts[1])
                        data['ytd']['total_rooms'] = int(parts[3])

                if line.strip().startswith('Out Of Order'):
                    parts = re.findall(r'\d+', line)
                    if len(parts) >= 5:
                        data['for_day']['ooo_rooms'] = int(parts[0])
                        data['mtd']['ooo_rooms'] = int(parts[1])
                        data['ytd']['ooo_rooms'] = int(parts[3])

                if line.strip().startswith('Comp Rooms') and 'Total' not in line:
                    parts = re.findall(r'\d+', line)
                    if len(parts) >= 5:
                        data['for_day']['comp_rooms'] = int(parts[0])
                        data['mtd']['comp_rooms'] = int(parts[1])
                        data['ytd']['comp_rooms'] = int(parts[3])

                if 'Total Occupied Rooms' in line and 'ADR' not in line:
                    parts = re.findall(r'\d+', line)
                    if len(parts) >= 5:
                        data['for_day']['rooms_sold'] = int(parts[0])
                        data['mtd']['rooms_sold'] = int(parts[1])
                        data['ytd']['rooms_sold'] = int(parts[3])

                if 'ADR for Total Occupied Rooms' in line:
                    parts = re.findall(r'\d+\.?\d*', line)
                    if len(parts) >= 5:
                        data['for_day']['adr'] = round(float(parts[0]), 2)
                        data['mtd']['adr'] = round(float(parts[1]), 2)
                        data['ytd']['adr'] = round(float(parts[3]), 2)

                if line.strip().startswith('RevPar') and 'STR' not in line:
                    parts = re.findall(r'\d+\.?\d*', line)
                    if len(parts) >= 5:
                        data['for_day']['revpar'] = round(float(parts[0]), 2)
                        data['mtd']['revpar'] = round(float(parts[1]), 2)
                        data['ytd']['revpar'] = round(float(parts[3]), 2)

                if 'Occ% of Total Rooms' in line and 'STR' not in line:
                    parts = re.findall(r'\d+\.?\d*', line)
                    if len(parts) >= 5:
                        data['for_day']['occp_pct'] = round(float(parts[0]), 2)
                        data['mtd']['occp_pct'] = round(float(parts[1]), 2)
                        data['ytd']['occp_pct'] = round(float(parts[3]), 2)

                if 'Total Room Revenue' in line:
                    parts = re.findall(r'-?\d+[\d,]*\.?\d*', line)
                    if len(parts) >= 5:
                        data['for_day']['room_revenue'] = round(abs(float(parts[0].replace(',', ''))), 2)
                        data['mtd']['room_revenue'] = round(float(parts[1].replace(',', '')), 2)
                        data['ytd']['room_revenue'] = round(float(parts[3].replace(',', '')), 2)

                if line.strip().startswith('Other Revenue') and 'Total' not in line:
                    parts = re.findall(r'-?\d+[\d,]*\.?\d*', line)
                    if len(parts) >= 5:
                        data['for_day']['other_revenue'] = round(abs(float(parts[0].replace(',', ''))), 2)
                        data['mtd']['other_revenue'] = round(float(parts[1].replace(',', '')), 2)
                        data['ytd']['other_revenue'] = round(float(parts[3].replace(',', '')), 2)

                if line.strip().startswith('Total Revenue') and 'Room' not in line:
                    parts = re.findall(r'-?\d+[\d,]*\.?\d*', line)
                    if len(parts) >= 5:
                        data['for_day']['total_revenue'] = round(abs(float(parts[0].replace(',', ''))), 2)
                        data['mtd']['total_revenue'] = round(float(parts[1].replace(',', '')), 2)
                        data['ytd']['total_revenue'] = round(float(parts[3].replace(',', '')), 2)

        return data

    def find_and_process_all_pdfs(self):
        """Automatically find and process all PDFs in the folder"""
        print("\n" + "="*70)
        print("AUTO PDF PROCESSOR - READING ALL PDFs IN FOLDER")
        print("="*70)
        print(f"\nFolder: {self.folder_path}\n")

        # Find all PDF files in the folder
        pdf_files = list(self.folder_path.glob("*.pdf"))

        if not pdf_files:
            print("❌ No PDF files found in the folder!")
            return

        print(f"✓ Found {len(pdf_files)} PDF file(s):\n")
        for pdf in pdf_files:
            print(f"  - {pdf.name}")

        print("\n" + "-"*70)
        print("PROCESSING PDFs...")
        print("-"*70 + "\n")

        # Process each PDF
        for pdf_path in pdf_files:
            try:
                pdf_name_lower = pdf_path.name.lower()

                if 'candlewood' in pdf_name_lower or 'burlington' in pdf_name_lower:
                    print(f"📄 {pdf_path.name}")
                    data = self.extract_candlewood_data(pdf_path)
                    print(f"   ✓ Extracted: {data['name']}")
                    self.hotels_data.append(data)

                elif 'tps' in pdf_name_lower or 'niagara' in pdf_name_lower:
                    print(f"📄 {pdf_path.name}")
                    data = self.extract_tps_niagara_data(pdf_path)
                    print(f"   ✓ Extracted: {data['name']}")
                    self.hotels_data.append(data)

                elif 'bayview' in pdf_name_lower or 'wildwood' in pdf_name_lower:
                    print(f"📄 {pdf_path.name}")
                    data = self.extract_bayview_data(pdf_path)
                    print(f"   ✓ Extracted: {data['name']}")
                    self.hotels_data.append(data)

                else:
                    print(f"⚠️  {pdf_path.name} - Unknown format, skipped")

            except Exception as e:
                print(f"❌ Error processing {pdf_path.name}: {str(e)}")

        # Create Excel report
        if self.hotels_data:
            print("\n" + "="*70)
            print("CREATING EXCEL REPORT...")
            print("="*70 + "\n")
            self.create_excel_report()
        else:
            print("\n❌ No valid data extracted from PDFs")

    def create_excel_report(self):
        """Create Excel file with all data"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Report"

        # Styling
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        title_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
        green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set column widths
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18

        # Headers
        ws['A1'] = 'S. No.'
        ws['B1'] = '1'
        ws['C1'] = '2'
        ws['D1'] = '3'

        for cell in ['A1', 'B1', 'C1', 'D1']:
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            ws[cell].fill = title_fill
            ws[cell].border = thin_border

        ws['B2'] = 'Ascend Collection'
        ws.merge_cells('B2:D2')
        ws['B2'].fill = header_fill
        ws['B2'].font = Font(bold=True, color="0000FF")
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

        ws['A2'] = 'Particulars'
        ws['A2'].font = Font(bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A2'].fill = title_fill

        ws['A3'] = ''
        ws['B3'] = 'A'
        ws['C3'] = 'B'
        ws['D3'] = 'C'

        for cell in ['B3', 'C3', 'D3']:
            ws[cell].font = Font(bold=True)
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            ws[cell].fill = title_fill
            ws[cell].border = thin_border

        # Metrics list
        metrics = [
            ('Total Rooms', 'total_rooms'),
            ('Rooms Sold', 'rooms_sold'),
            ('Comp Rooms', 'comp_rooms'),
            ('OOO Rooms', 'ooo_rooms'),
            ('ADR', 'adr'),
            ('RevPar', 'revpar'),
            ('Occp%', 'occp_pct'),
            ('Room Revenue', 'room_revenue'),
            ('F & B Revenue', 'fb_revenue'),
            ('Other Revenue', 'other_revenue'),
            ('Total Revenue', 'total_revenue')
        ]

        current_row = 5

        # FOR THE DAY
        ws[f'A{current_row}'] = 'For the Day'
        ws[f'A{current_row}'].font = Font(bold=True)
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].fill = yellow_fill
            ws[f'{col}{current_row}'].border = thin_border

        for idx, (label, key) in enumerate(metrics):
            row = current_row + 1 + idx
            ws[f'A{row}'] = label
            ws[f'A{row}'].border = thin_border

            for col_idx, hotel_data in enumerate(self.hotels_data[:3]):
                col = chr(66 + col_idx)
                value = hotel_data['for_day'].get(key, '')
                ws[f'{col}{row}'] = value if value else ''
                ws[f'{col}{row}'].border = thin_border
                ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

        current_row += len(metrics) + 2

        # MTD
        ws[f'A{current_row}'] = 'MTD'
        ws[f'A{current_row}'].font = Font(bold=True)
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].fill = green_fill
            ws[f'{col}{current_row}'].border = thin_border

        for idx, (label, key) in enumerate(metrics):
            row = current_row + 1 + idx
            ws[f'A{row}'] = label
            ws[f'A{row}'].border = thin_border

            for col_idx, hotel_data in enumerate(self.hotels_data[:3]):
                col = chr(66 + col_idx)
                value = hotel_data['mtd'].get(key, '')
                ws[f'{col}{row}'] = value if value else ''
                ws[f'{col}{row}'].border = thin_border
                ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

        current_row += len(metrics) + 2

        # YTD
        ws[f'A{current_row}'] = 'YTD'
        ws[f'A{current_row}'].font = Font(bold=True)
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].fill = blue_fill
            ws[f'{col}{current_row}'].border = thin_border

        for idx, (label, key) in enumerate(metrics):
            row = current_row + 1 + idx
            ws[f'A{row}'] = label
            ws[f'A{row}'].border = thin_border

            for col_idx, hotel_data in enumerate(self.hotels_data[:3]):
                col = chr(66 + col_idx)
                value = hotel_data['ytd'].get(key, '')
                ws[f'{col}{row}'] = value if value else ''
                ws[f'{col}{row}'].border = thin_border
                ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

        # Add hotel names
        for col_idx, hotel_data in enumerate(self.hotels_data[:3]):
            col = chr(66 + col_idx)
            ws[f'{col}3'] = hotel_data['name']
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')

        # Save
        output_path = self.folder_path / 'Daily_Revenue_Report_Hotel.xlsx'
        wb.save(output_path)

        print(f"✅ Excel file created: {output_path.name}")
        print(f"✅ Hotels processed: {len(self.hotels_data)}")
        print("\n" + "="*70)
        print("✅ DONE! You can now open the Excel file!")
        print("="*70 + "\n")


# RUN THE CONVERTER
if __name__ == "__main__":
    # Folder containing PDFs - change this to your folder path
    FOLDER_PATH = "/Users/gupta/Downloads/hotel-data/"

    # Create converter and process all PDFs in folder
    converter = AutoHotelPDFConverter(FOLDER_PATH)
    converter.find_and_process_all_pdfs()